import warnings
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from pygam import *
import statsmodels.formula.api as smf
from statsmodels.tsa.seasonal import STL
from plotly.subplots import make_subplots
import nbformat
import requests
import io
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, NamedStyle, Border, Side

warnings.simplefilter(action = "ignore")

database_revenue = pd.read_excel("data/Cafeteria Fictícia - Planilhas Unificadas.xlsx", sheet_name = "Receita")

column_map = {
	"Data": "date",
    "Preço": "price",
    "Nome do Café": "item"
}

valid_columns = []
for column_name in database_revenue.columns:
    if column_name in column_map.keys():
        valid_columns.append(column_name)

database_revenue = database_revenue[valid_columns]
database_revenue.rename(columns = column_map, inplace = True)

# database_revenue.head(10)

database_expense = pd.read_excel("data/Cafeteria Fictícia - Planilhas Unificadas.xlsx", sheet_name = "Despesa")

column_map = {
	"Data": "date",
    "Insumo": "material",
    "Quantidade Adquirida": "quantity_purchased",
    "Custo Unitário": "unit_cost",
    "Subtotal": "subtotal"
}

valid_columns = []
for column_name in database_expense.columns:
    if column_name in column_map.keys():
        valid_columns.append(column_name)

database_expense = database_expense[valid_columns]
database_expense.rename(columns = column_map, inplace = True)

# database_expense.head(10)

database_balance_accounts = pd.read_excel("data/Cafeteria Fictícia - Planilhas Unificadas.xlsx", sheet_name = "Balanço Patrimonial")

column_map = {
   "Rubrica": "heading",
   "1T 2024": "1T 2024",
   "2T 2024": "2T 2024",
   "3T 2024": "3T 2024",
   "4T 2024": "4T 2024",
   "1T 2025": "1T 2025",
   "2T 2025": "2T 2025",
   "3T 2025": "3T 2025",
   "4T 2025": "4T 2025",
}

valid_columns = []
for column_name in database_balance_accounts.columns:
    if column_name in column_map.keys():
        valid_columns.append(column_name)

database_balance_accounts = database_balance_accounts[valid_columns]
database_balance_accounts.rename(columns = column_map, inplace = True)

# database_balance_accounts.head(10)

database_employees = pd.read_excel("data/Cafeteria Fictícia - Planilhas Unificadas.xlsx", sheet_name = "Quadro de Funcionários")

column_map = {
	"Data": "date",
    "Funcionário": "employee",
    "Cargo": "position",
    "Salário": "wage"
}

valid_columns = []
for column_name in database_employees.columns:
    if column_name in column_map.keys():
        valid_columns.append(column_name)

database_employees = database_employees[valid_columns]
database_employees.rename(columns = column_map, inplace = True)

# database_employees.head(10)

sales_summary = database_revenue.groupby(["item", "price"]).size().reset_index(name = "quantity_sold")

figure1 = px.scatter(
    sales_summary, x = "quantity_sold", y = "price", color = "item", trendline = "ols",
    labels = {"price": "Preço (R$)", "quantity_sold": "Quantidade Vendida", "item": "Item"},
    title = "Análise Exploratória — Demandas Inversas", width = 1000, height = 500
)

figure1.update_layout(
    title_font_size = 18, font = dict(size = 14, family = "Arial", color = "black"),
    plot_bgcolor = "white", paper_bgcolor = "white",
    legend = dict(title = "", borderwidth = 0, font_size = 12, bgcolor = "rgba(0,0,0,0)"),
    xaxis = dict(showgrid = True, gridcolor = "lightgrey", zeroline = False, title_font_size = 14),
    yaxis = dict(showgrid = True, gridcolor = "lightgrey", zeroline = False, title_font_size = 14)
)

# figure1.show()

figure2 = px.violin(
    database_revenue, x = "item",  y = "price", color = "item", box = False, points = "all",
    labels = {"price": "Preço (R$)", "item": "Item"},
    title = "Análise Exploratória — Distribuições de Preços", width = 1000, height = 500
)

figure2.update_layout(
    title_font_size = 18, font = dict(size = 14, family = "Arial", color = "black"),
    plot_bgcolor = "white", paper_bgcolor = "white",
    showlegend = False,
    xaxis = dict(showgrid = True, gridcolor = "lightgrey", zeroline = False, title_font_size = 14),
    yaxis = dict(showgrid = True, gridcolor = "lightgrey", zeroline = False, title_font_size = 14)
)

# figure2.show()

elasticities = []

latest_prices = database_revenue.sort_values("date").groupby("item").tail(1)
latest_prices = latest_prices.set_index("item")["price"]

for item in sales_summary["item"].unique():
    item_data = sales_summary[sales_summary["item"] == item]

    item_data["quantity_sold"] = np.log(item_data["quantity_sold"])
    item_data["price"] = np.log(item_data["price"])

    log_log = smf.ols("quantity_sold ~ price", data = item_data).fit()
    beta_0, beta_1 = log_log.params

    P0 = latest_prices[item]
    Q0 = beta_0 + beta_1 * P0

    current_elasticity = beta_1 * (P0 / Q0)

    elasticities.append({
        "item": item,
        "current_price": P0,
        "predicted_quantity_sold": Q0,
        "current_elasticity": np.abs(current_elasticity)
    })

elasticities = pd.DataFrame(elasticities)

figure3 = px.bar(
    elasticities,
    x = "item", y = "current_elasticity", color = "item",
    labels = {"item": "Item", "current_elasticity": "Nível"},
    title = "Análise Exploratória — Elasticidades-preço da Demanda Atuais", width = 1000, height = 500
)

for index, row in elasticities.iterrows():
    figure3.add_annotation(
        x = row["item"], y = row["current_elasticity"],
        text = f"<b>{row["current_elasticity"]:.2f}</b>".replace(".", ","),
        showarrow = False, font = dict(color = "white", size = 12),
        align = "center", bordercolor = "black",
        borderwidth = 1, bgcolor = "black", opacity = 0.8
    )

figure3.update_layout(
    title_font_size = 18,
    font = dict(size = 14, family = "Arial", color = "black"),
    plot_bgcolor = "white",
    paper_bgcolor = "white",
    legend = dict(title = "", borderwidth = 0, font_size = 12, bgcolor = "rgba(0,0,0,0)"),
    xaxis = dict(showgrid = True, gridcolor = "lightgrey", zeroline = False, title_font_size = 14),
    yaxis = dict(showgrid = True, gridcolor = "lightgrey", zeroline = False, title_font_size = 14),
    separators = ",."
)

# figure3.show()

accumulated_revenue = (database_revenue.groupby(["date", "item"])["price"]
                 .sum().groupby(level = 1 ).cumsum().reset_index(name = "accumulated_revenue"))

accumulated_revenue["accumulated_revenue"] /= 1000

figure4 = px.line(accumulated_revenue, x = "date", y = "accumulated_revenue", color = "item",
               labels = {"date": "Data", "accumulated_revenue": "Receita acumulada (mil R$)", "item": "Item"},
               title = "Análise Exploratória — Receitas Acumuladas", width = 1000, height = 500)

figure4.update_layout(
    title_font_size = 18, font = dict(size = 14, family = "Arial", color = "black"),
    plot_bgcolor = "white", paper_bgcolor = "white",
    legend = dict(title = "", font_size = 12, bgcolor = "rgba(0,0,0,0)"),
    xaxis = dict(showgrid = True, gridcolor = "lightgrey", zeroline = False, title_font_size = 14, tickformat = "%d/%m/%Y"),
    yaxis = dict(showgrid = True, gridcolor = "lightgrey", zeroline = False, title_font_size = 14)
)

# figure4.show()

daily_revenue = (database_revenue.groupby(["date", "item"])["price"]
                 .sum().reset_index(name = "daily_revenue"))

figure5 = px.line(
    daily_revenue, x = "date", y = "daily_revenue", color = "item",
    labels = {"date": "Data", "daily_revenue": "Receita diária (R$)", "item": "Item"},
    title = "Análise Exploratória — Receitas Diárias", width = 1000, height = 500
)

figure5.update_layout(
    title_font_size = 18, font = dict(size = 14, family = "Arial", color = "black"),
    plot_bgcolor = "white", paper_bgcolor = "white",
    legend = dict(title = "", font_size = 12, bgcolor = "rgba(0,0,0,0)"),
    xaxis = dict(showgrid = True, gridcolor = "lightgrey", zeroline = False, title_font_size = 14, tickformat = "%d/%m/%Y"),
    yaxis = dict(showgrid = True, gridcolor = "lightgrey", zeroline = False, title_font_size = 14)
)

# figure5.show()

database_revenue["date"] = pd.to_datetime(database_revenue["date"])

translated_weekdays = {0: "Segunda-feira", 1: "Terça-feira", 2: "Quarta-feira", 3: "Quinta-feira", 4: "Sexta-feira", 5: "Sábado", 6: "Domingo"}
database_revenue["weekday"] = database_revenue["date"].dt.dayofweek.map(translated_weekdays)

weekdays_revenue = (database_revenue.groupby(["date", "weekday"])["price"]
                   .mean().reset_index(name = "weekdays_revenue"))

weekdays_revenue["weekday"] = pd.Categorical(weekdays_revenue["weekday"], ordered = True)

figure6 = px.line(
    weekdays_revenue, x = "date", y = "weekdays_revenue", color = "weekday",
    labels = {"date": "Data", "weekdays_revenue": "Receita média (R$)", "weekday": "Dia da semana"},
    title = "Análise Exploratória — Receitas Médias por Dia da Semana", width = 1000, height = 500
)

figure6.update_layout(
    title_font_size = 18,
    font = dict(size = 14, family = "Arial", color = "black"),
    plot_bgcolor = "white",
    paper_bgcolor = "white",
    legend = dict(title = "", font_size = 12, bgcolor = "rgba(0,0,0,0)"),
    xaxis = dict(showgrid = True, gridcolor = "lightgrey", zeroline = False, title_font_size = 14, tickformat = "%d/%m/%Y"),
    yaxis = dict(showgrid = True, gridcolor = "lightgrey", zeroline = False, title_font_size = 14)
)

# figure6.show()

daily_revenue = (database_revenue.groupby(["date", "item"])["price"]
                 .sum().reset_index(name="daily_revenue"))

daily_revenue["date"] = pd.to_datetime(daily_revenue["date"])
daily_revenue = daily_revenue.set_index("date")

weekly_revenue = (daily_revenue.groupby("item")
                  .resample("W")["daily_revenue"]
                  .mean()
                  .reset_index())

weekly_revenue["total_week"] = weekly_revenue.groupby("date")["daily_revenue"].transform("sum")
weekly_revenue["percentage_revenue"] = weekly_revenue["daily_revenue"] / weekly_revenue["total_week"] * 100

figure7 = px.area(
    weekly_revenue, 
    x = "date", y = "percentage_revenue", color = "item",
    labels = {"date": "Data", "percentage_revenue": "Participação", "item": "Item"},
    title = "Análise Exploratória — Composição Dinâmica da Receita", width = 1000, height = 500
)

figure7.update_layout(
    title_font_size = 18,
    font = dict(size = 14, family = "Arial", color = "black"),
    plot_bgcolor = "white",
    paper_bgcolor = "white",
    legend = dict(title = "", font_size = 12, bgcolor = "rgba(0,0,0,0)"),
    xaxis = dict(showgrid = True, gridcolor = "lightgrey", zeroline = False, title_font_size = 14, tickformat = "%d/%m/%Y"),
    yaxis = dict(showgrid = True, gridcolor = "lightgrey", zeroline = False, title_font_size = 14, ticksuffix = "%")
)

# figure7.show()

database_employees["date"] = pd.to_datetime(database_employees["date"], format = "%d/%m/%Y")
current_year = database_employees["date"].dt.year.max()
current_year = database_employees[database_employees["date"].dt.year == current_year]

employee_summary = current_year.groupby("position").agg(
    employee_count = ("employee", "nunique"),
    average_wage = ("wage", "mean")
).reset_index()

figure8 = px.bar(
    employee_summary,
    x = "position", y = "employee_count", color = "position",
    labels = {"position": "Cargo", "employee_count": "Número de Funcionários"},
    title = "Análise Exploratória — Funcionários por Cargo e Salário Médio", width = 1000, height = 500
)

for index, row in employee_summary.iterrows():
    figure8.add_annotation(
        x = row["position"], y = row["employee_count"],
        text = f"<b>R$ {row['average_wage']:,.2f}</b>".replace(",", "X").replace(".", ",").replace("X", "."),
        showarrow = False, font = dict(color = "white", size = 12),
        align = "center", bordercolor = "black",
        borderwidth = 1, bgcolor = "black", opacity = 0.8
    )

figure8.update_layout(
    title_font_size = 18,
    font = dict(size = 14, family = "Arial", color = "black"),
    plot_bgcolor = "white",
    paper_bgcolor = "white",
    legend = dict(title = "", borderwidth = 0, font_size = 12, bgcolor = "rgba(0,0,0,0)"),
    xaxis = dict(showgrid = True, gridcolor = "lightgrey", zeroline = False, title_font_size = 14),
    yaxis = dict(showgrid = True, gridcolor = "lightgrey", zeroline = False, title_font_size = 14),
    separators = ",."
)

# figure8.show()

optimal_prices = []
gam_results = {}

for item in sales_summary["item"].unique():
    item_data = sales_summary[sales_summary["item"] == item]

    price_values = item_data[["price"]].values
    quantity_sold_values = item_data["quantity_sold"].values

    gam = PoissonGAM(s(0, n_splines = 5, spline_order = 3, constraints = "monotonic_dec")).gridsearch(price_values, quantity_sold_values)
    price_range = np.linspace(price_values.min(), price_values.max(), 100)

    demand_estimated = gam.predict(price_range)
    revenue_estimated = price_range * demand_estimated / 1000

    optimal_index = np.argmax(revenue_estimated)
    optimal_price = price_range[optimal_index]
    optimal_quantity_sold = demand_estimated[optimal_index]

    optimal_prices.append({
        "item": item,
        "optimal_price": round(optimal_price, 2),
        "expected_quantity_sold": round(optimal_quantity_sold, 2),
        "expected_revenue": round(revenue_estimated[optimal_index], 2)
    })

    gam_results[item] = {
        "price_range": price_range,
        "demand_estimated": demand_estimated,
        "revenue_estimated": revenue_estimated,
        "optimal_price": optimal_price,
        "optimal_quantity_sold": optimal_quantity_sold
    }

figure9 = go.Figure()

colors = ["#636efa", "#ef553b", "#00cc96", "#ab63fa", "#ffa15a", 
          "#19d3f3", "#ff6692", "#b6e880", "#bcbd22", "#17becf"]

for index, item in enumerate(sales_summary["item"].unique()):
    item_data = sales_summary[sales_summary["item"] == item]
    result = gam_results[item]
    color = colors[index % len(colors)]
    visible = (item == sales_summary["item"].unique()[0])

    figure9.add_trace(go.Scatter(
        x = item_data["price"], y = item_data["quantity_sold"], 
        mode = "markers", name = "Observado",
        marker = dict(size = 8, color = color, opacity = 0.6),
        visible = visible, legendgroup = "observed", showlegend = True
    ))

    figure9.add_trace(go.Scatter(
        x = result["price_range"], y = result["demand_estimated"],
        mode = "lines", name = "Demanda estimada",
        line = dict(color = color, width = 2),
        visible = visible, legendgroup = "demand", showlegend = True
    ))

    figure9.add_trace(go.Scatter(
        x = result["price_range"], y = result["revenue_estimated"],
        mode = "lines", name = "Receita",
        line = dict(color = color, dash = "dot", width = 2),
        yaxis = "y2", visible = visible, 
        legendgroup = "revenue", showlegend = True
    ))

    figure9.add_trace(go.Scatter(
        x = [result["optimal_price"]], y = [result["optimal_quantity_sold"]],
        mode = "markers+text", text = [f"Ótimo: R$ {result["optimal_price"]:.2f}"],
        textposition = "top center", marker = dict(color = color, size = 10),
        name = "Ótimo", visible = visible, 
        legendgroup = "optimal", showlegend = True
    ))

buttons = []

for item in sales_summary["item"].unique():
    buttons.append({
        "label": item, "method": "update",
        "args": [{"visible": [item == coffee for coffee in sales_summary["item"].unique() for _ in range(4)],
                  "title": f"Forecasting e Relacionados — Generalized Additive Model (GAM)"}]
    })

figure9.update_layout(
    title = f"Forecasting e Relacionados — Generalized Additive Model (GAM)",
    title_font_size = 18, font = dict(size = 14, family = "Arial", color = "black"),
    width = 1000, height = 500, plot_bgcolor = "white", paper_bgcolor = "white",
    xaxis = dict(title = "Preço (R$)", showgrid = True, gridcolor = "lightgrey", 
                 zeroline = False, title_font_size = 14),
    yaxis = dict(title = "Quantidade Vendida", showgrid = True, gridcolor = "lightgrey", 
                 zeroline = False, title_font_size = 14),
    yaxis2 = dict(title = "Receita (mil R$)", overlaying = "y", side = "right",
                  showgrid = False, zeroline = False, title_font_size = 14),
    legend = dict(title = "", borderwidth = 0, font_size = 12, 
                  bgcolor = "rgba(0,0,0,0)", orientation = "v", x = 1.08, y = 1),
    updatemenus = [dict(
        buttons = buttons, direction = "down", showactive = True,
        x = 1.0, xanchor = "right", y = 1.15, yanchor = "top"
    )]
)

# figure9.show()

optimal_elasticities = []

for item in sales_summary["item"].unique():
    result = gam_results[item]
    price_range = result["price_range"]
    demand_estimated = result["demand_estimated"]
    optimal_index = np.argmax(result["revenue_estimated"])

    optimal_P = price_range[optimal_index]
    optimal_Q = demand_estimated[optimal_index]

    if optimal_index > 0 and optimal_index < len(price_range) - 1:
        dP = price_range[optimal_index + 1] - price_range[optimal_index - 1]
        dQ = demand_estimated[optimal_index + 1] - demand_estimated[optimal_index - 1]
        dQ_dP = dQ / dP
    elif optimal_index == 0:
        dP = price_range[1] - price_range[0]
        dQ = demand_estimated[1] - demand_estimated[0]
        dQ_dP = dQ / dP
    else:
        dP = price_range[-1] - price_range[-2]
        dQ = demand_estimated[-1] - demand_estimated[-2]
        dQ_dP = dQ / dP

    optimal_elasticity = dQ_dP * (optimal_P / optimal_Q)

    optimal_elasticities.append({
        "item": item,
        "optimal_price": optimal_P,
        "predicted_quantity_sold": optimal_Q,
        "optimal_elasticity": np.abs(optimal_elasticity)
    })

optimal_elasticities = pd.DataFrame(optimal_elasticities)

figure10 = px.bar(
    optimal_elasticities,
    x = "item", y = "optimal_elasticity", color = "item",
    labels = {"item": "Item", "optimal_elasticity": "Nível"},
    title = "Forecasting e Relacionados — Elasticidades-preço da Demanda Ótimas", width = 1000, height = 500
)

for index, row in optimal_elasticities.iterrows():
    figure10.add_annotation(
        x = row["item"], y = row["optimal_elasticity"],
        text = f"<b>{row["optimal_elasticity"]:.2f}</b>".replace(".", ","),
        showarrow = False, font = dict(color = "white", size = 12),
        align = "center", bordercolor = "black",
        borderwidth = 1, bgcolor = "black", opacity = 0.8
    )

figure10.update_layout(
    title_font_size = 18,
    font = dict(size = 14, family = "Arial", color = "black"),
    plot_bgcolor = "white",
    paper_bgcolor = "white",
    legend = dict(title = "", borderwidth = 0, font_size = 12, bgcolor = "rgba(0,0,0,0)"),
    xaxis = dict(showgrid = True, gridcolor = "lightgrey", zeroline = False, title_font_size = 14),
    yaxis = dict(showgrid = True, gridcolor = "lightgrey", zeroline = False, title_font_size = 14),
    separators = ",."
)

# figure10.show()

daily_revenue = (database_revenue.groupby(["date", "item"])["price"]
                 .sum().reset_index(name = "daily_revenue"))

daily_revenue["date"] = pd.to_datetime(daily_revenue["date"])

items_list = list(sales_summary["item"].unique())

decomposition_frames = []

for index, item in enumerate(items_list):
    series = (daily_revenue.loc[daily_revenue["item"] == item, ["date", "daily_revenue"]]
                                  .set_index("date")
                                  .sort_index()
                                  .asfreq("D"))

    series["daily_revenue"] = series["daily_revenue"].fillna(0.0)

    stl = STL(series["daily_revenue"], period = 7, robust = True)
    stl_result = stl.fit()

    decomposition_frame = pd.DataFrame({
        "date": series.index,
        "item": item,
        "trend": stl_result.trend,
        "seasonal": stl_result.seasonal,
        "residual": stl_result.resid
    }).reset_index(drop = True)

    decomposition_frames.append(decomposition_frame)

decomposition_data = pd.concat(decomposition_frames, ignore_index = True)

figure11 = make_subplots(
    rows = 3, cols = 1, shared_xaxes = True, vertical_spacing = 0.08,
    subplot_titles = ("Tendência", "Sazonalidade", "Resíduo")
)

trace_visibility = []

for index, item in enumerate(items_list):
    color = colors[index % len(colors)]
    slice = decomposition_data[decomposition_data["item"] == item]

    is_visible = (index == 0)

    figure11.add_trace(
        go.Scatter(
            x = slice["date"], y = slice["trend"],
            mode = "lines", name = "Tendência",
            line = dict(width = 2, color = color),
            visible = is_visible,
            showlegend = True
        ),
        row = 1, col = 1
    )
    trace_visibility.append(is_visible)

    figure11.add_trace(
        go.Scatter(
            x = slice["date"], y = slice["seasonal"],
            mode = "lines", name = "Sazonalidade",
            line = dict(width = 2.75, color = color),
            visible = is_visible,
            showlegend = True
        ),
        row = 2, col = 1
    )
    trace_visibility.append(is_visible)

    figure11.add_trace(
        go.Scatter(
            x = slice["date"], y = slice["residual"],
            mode = "lines", name = "Resíduo",
            line = dict(width = 2, color = color, dash = "dot"),
            visible = is_visible,
            showlegend = True
        ),
        row = 3, col = 1
    )
    trace_visibility.append(is_visible)

buttons = []
traces_per_item = 3
total_traces = traces_per_item * len(items_list)

for index, item in enumerate(items_list):
    visibility_mask = [False] * total_traces
    start = index * traces_per_item
    for k in range(traces_per_item):
        visibility_mask[start + k] = True

    buttons.append(dict(
        label = item,
        method = "update",
        args = [
            {"visible": visibility_mask},
            {"title": f"Decomposição de Receita — {item}"}
        ]
    ))

figure11.update_layout(
    title = f"Forecasting e Relacionados — Tendência, Sazonalidade e Resíduo",
    title_font_size = 18,
    font = dict(size = 14, family = "Arial", color = "black"),
    width = 1000, height = 700,
    plot_bgcolor = "white", paper_bgcolor = "white",
    legend = dict(title = "", borderwidth = 0, font_size = 12, bgcolor = "rgba(0,0,0,0)"),
    updatemenus = [dict(
        buttons = buttons, direction = "down", showactive = True,
        x = 1.0, xanchor = "right", y = 1.15, yanchor = "top"
    )]
)

figure11.update_xaxes(
    showgrid = True, gridcolor = "lightgrey", zeroline = False, title_font_size = 14,
    tickformat = "%d/%m/%Y", row = 3, col = 1, title_text = "Data"
)
figure11.update_xaxes(
    showgrid = True, gridcolor = "lightgrey", zeroline = False, tickformat = "%d/%m/%Y", row = 1, col = 1
)
figure11.update_xaxes(
    showgrid = True, gridcolor = "lightgrey", zeroline = False, tickformat = "%d/%m/%Y", row = 2, col = 1
)

figure11.update_yaxes(title_text = "Nível", showgrid = True, gridcolor = "lightgrey",
                     zeroline = False, title_font_size = 14, row = 1, col = 1)
figure11.update_yaxes(title_text = "Nível", showgrid = True, gridcolor = "lightgrey",
                     zeroline = False, title_font_size = 14, row = 2, col = 1)
figure11.update_yaxes(title_text = "Nível", showgrid = True, gridcolor = "lightgrey",
                     zeroline = False, title_font_size = 14, row = 3, col = 1)

# figure11.show()

comparison_table = pd.DataFrame(optimal_prices)

comparison_table["current_price"] = comparison_table["item"].map(latest_prices)
comparison_table["percent_difference"] = (comparison_table["optimal_price"] - comparison_table["current_price"]) / comparison_table["current_price"] * 100
comparison_table["estimated_revenue"] = comparison_table["optimal_price"] * comparison_table["expected_quantity_sold"]

comparison_table = comparison_table[[
    "item",
    "current_price",
    "optimal_price",
    "percent_difference",
    "expected_quantity_sold",
    "estimated_revenue"
]]

comparison_table.columns = [
    "Item",
    "Preço Atual (R$)",
    "Preço Ótimo (R$)",
    "Diferença (%)",
    "Quantidade Vendida Estimada",
    "Receita Estimada (R$)"
]

for column in comparison_table.columns:
    if comparison_table[column].dtype in ["float64", "int64"]:
        comparison_table[column] = comparison_table[column].apply(lambda x: f"{x:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))

comparison_table.to_excel("Entregável - Tabela de Comparação.xlsx", index = False)

revision = pd.read_excel("Entregável - Tabela de Comparação.xlsx")
revision.head(5).style.hide(axis = "index")

database_revenue["date"] = pd.to_datetime(database_revenue["date"])
database_expense["date"] = pd.to_datetime(database_expense["date"])

monthly_revenue = database_revenue.groupby(pd.Grouper(key = "date", freq = "M"))["price"].sum().reset_index()
monthly_revenue.rename(columns = {"price": "monthly_revenue"}, inplace = True)

monthly_expense = database_expense.groupby(pd.Grouper(key = "date", freq = "M"))["subtotal"].sum().reset_index()
monthly_expense.rename(columns = {"subtotal": "monthly_expense"}, inplace = True)

date_range = pd.date_range(start = monthly_revenue["date"].min(), end = monthly_revenue["date"].max(), freq = "M")
calendar = pd.DataFrame({"date": date_range})

cash_flow = calendar.merge(monthly_revenue, on = "date", how = "left")
cash_flow = cash_flow.merge(monthly_expense, on = "date", how = "left")
cash_flow.fillna(0, inplace = True)

cash_flow["net_income"] = cash_flow["monthly_revenue"] - cash_flow["monthly_expense"]
cash_flow["net_margin_percentage"] = (cash_flow["net_income"] / cash_flow["monthly_revenue"]) * 100
cash_flow["net_margin_percentage"].replace([np.inf, -np.inf], 0, inplace = True)

cash_flow["monthly_revenue"] = cash_flow["monthly_revenue"] / 1000
cash_flow["monthly_expense"] = cash_flow["monthly_expense"] / 1000

figure12 = go.Figure()

figure12.add_trace(go.Bar(
    x = cash_flow["date"],
    y = cash_flow["monthly_revenue"],
    name = "Entradas",
    marker_color = "#00cc96",
))

figure12.add_trace(go.Bar(
    x = cash_flow["date"],
    y = -cash_flow["monthly_expense"],
    name = "Saídas",
    marker_color = "#ef553b",
))

figure12.add_trace(go.Scatter(
    x = cash_flow["date"],
    y = cash_flow["net_margin_percentage"],
    name = "Margem Líquida",
    line = dict(color = "darkgray", width = 3, dash = "dot"),
    yaxis = "y2"
))

for index, row in cash_flow.iterrows():
    figure12.add_annotation(
        x = row["date"], y = row["monthly_revenue"],
        text = f"<b>{row['monthly_revenue']:,.2f}</b>".replace(".", ","),
        showarrow = False, font = dict(color = "white", size = 12),
        align = "center", bordercolor = "black",
        borderwidth = 1, bgcolor = "black", opacity = 0.8
    )

for index, row in cash_flow.iterrows():
    figure12.add_annotation(
        x = row["date"], y = -row["monthly_expense"],
        text = f"<b>{row['monthly_expense']:,.2f}</b>".replace(".", ","),
        showarrow = False, font = dict(color = "white", size = 12),
        align = "center", bordercolor = "black",
        borderwidth = 1, bgcolor = "black", opacity = 0.8
    )

figure12.update_layout(
    title = "Fluxo de Caixa e Estoque — Movimentações Mensais",
    xaxis_title = "Data",
    yaxis_title = "Movimentação (R$ mil)",
    yaxis2 = dict(
        title = "Margem Líquida (%)",
        overlaying = "y",
        side = "right",
        range = [cash_flow["net_margin_percentage"].min() - 10, cash_flow["net_margin_percentage"].max() + 10],
        title_font_size = 14
    ),
    title_font_size = 18,
    font = dict(size = 14, family = "Arial", color = "black"),
    plot_bgcolor = "white",
    paper_bgcolor = "white",
    legend = dict(title = "", font_size = 12, bgcolor = "rgba(0,0,0,0)",
                  x = 1.1, y = 1, xanchor = "left", yanchor = "top"),
    xaxis = dict(
        showgrid = True, 
        gridcolor = "lightgrey", 
        zeroline = False, 
        title_font_size = 14,
        tickformat = "%m/%Y",
        dtick = "M1"
    ),
    yaxis = dict(
        showgrid = True, 
        gridcolor = "lightgrey", 
        zeroline = False, 
        title_font_size = 14
    ),
    barmode = "relative",
    width = 1000, height = 500
)

# figure12.show()

annual_cash_flow = cash_flow.copy()
annual_cash_flow.set_index('date', inplace=True)
annual_cash_flow['year'] = annual_cash_flow.index.year
annual_cash_flow = annual_cash_flow.groupby('year').agg({
    'monthly_expense': 'sum',
    'monthly_revenue': 'sum'
}).reset_index()
annual_cash_flow.rename(columns={'monthly_expense': 'annual_expense', 'monthly_revenue': 'annual_revenue'}, inplace=True)

last_year_cashflow = annual_cash_flow['year'].max()
projection_range = range(last_year_cashflow + 1, last_year_cashflow + 6)

last_revenue = annual_cash_flow[annual_cash_flow['year'] == last_year_cashflow]['annual_revenue'].values[0]
revenue_projection = []

# Revenue Projection
actual_revenue = last_revenue
for year in projection_range:
    actual_revenue *= 1.10 # Adjustable
    revenue_projection.append({
        'year': year,
        'Receitas': actual_revenue
})
    
revenue_projection = pd.DataFrame(revenue_projection)

projected_cash_flow_revenue = annual_cash_flow[['year', 'annual_revenue']].copy()
projected_cash_flow_revenue.rename(columns={'annual_revenue': 'Receitas'}, inplace=True)
projected_cash_flow_revenue = pd.concat([
    projected_cash_flow_revenue,
    revenue_projection
], ignore_index=True)


# Expense and Margin Projection
last_expense = annual_cash_flow[annual_cash_flow['year'] == last_year_cashflow]['annual_expense'].values[0]

expense_projection = []
actual_expense = last_expense
for year in projection_range:
    actual_expense *= 1.10 # Adjustable
    expense_projection.append({
        'year': year,
        'Despesas': actual_expense
})

expense_projection = pd.DataFrame(expense_projection)

projected_cash_flow_expense = annual_cash_flow[['year', 'annual_expense']].copy()
projected_cash_flow_expense.rename(columns={'annual_expense': 'Despesas'}, inplace=True)
projected_cash_flow_expense = pd.concat([
    projected_cash_flow_expense,
    expense_projection
], ignore_index=True)

# Projected cash flow
projected_cash_flow = pd.merge(projected_cash_flow_revenue, projected_cash_flow_expense, on='year', how='inner')

# Projected Margin
projected_cash_flow['Margem'] = projected_cash_flow['Receitas'] - projected_cash_flow['Despesas']

projected_cash_flow.rename(columns={'year': 'Ano'}, inplace=True)
projected_cash_flow = projected_cash_flow.set_index('Ano')
projected_cash_flow = projected_cash_flow.T
projected_cash_flow.index.name = None

#print(projected_cash_flow)

database_balance_accounts.set_index("heading", inplace = True)
database_balance_accounts = database_balance_accounts.apply(pd.to_numeric, errors = "coerce")

quarters = database_balance_accounts.columns.tolist()

current_liquidity = []
quick_liquidity = []
immediate_liquidity = []

for quarter in quarters:
    current_assets = database_balance_accounts.loc["Ativo Circulante", quarter]
    inventory = database_balance_accounts.loc["Estoque", quarter]
    cash_equivalents = database_balance_accounts.loc["Caixa e Equivalentes de Caixa", quarter]
    current_liabilities = database_balance_accounts.loc["Passivo Circulante", quarter]

    liquidity_current = current_assets / current_liabilities
    liquidity_quick = (current_assets - inventory) / current_liabilities
    liquidity_immediate = cash_equivalents / current_liabilities

    current_liquidity.append(liquidity_current)
    quick_liquidity.append(liquidity_quick)
    immediate_liquidity.append(liquidity_immediate)

liquidity_ratios = pd.DataFrame({
    "Trimestre": quarters,
    "Liquidez Corrente": current_liquidity,
    "Liquidez Seca": quick_liquidity,
    "Liquidez Imediata": immediate_liquidity
})

figure13 = go.Figure()

figure13.add_trace(go.Scatter(
    x = liquidity_ratios["Trimestre"], y = liquidity_ratios["Liquidez Corrente"],
    mode = "lines+markers", name = "Liquidez Corrente",
    line = dict(width = 2, color = "#636efa"), marker = dict(size = 8)
))

figure13.add_trace(go.Scatter(
    x = liquidity_ratios["Trimestre"], y = liquidity_ratios["Liquidez Seca"],
    mode = "lines+markers", name = "Liquidez Seca",
    line = dict(width = 2, color = "#ef553b"), marker = dict(size = 8)
))

figure13.add_trace(go.Scatter(
    x = liquidity_ratios["Trimestre"], y = liquidity_ratios["Liquidez Imediata"],
    mode = "lines+markers", name = "Liquidez Imediata",
    line = dict(width = 2, color = "#00cc96"), marker = dict(size = 8)
))

figure13.update_layout(
    title = "Fluxo de Caixa e Estoque — Indicadores de Liquidez",
    title_font_size = 18,
    font = dict(size = 14, family = "Arial", color = "black"),
    plot_bgcolor = "white",
    paper_bgcolor = "white",
    legend = dict(title = "", font_size = 12, bgcolor = "rgba(0,0,0,0)"),
    xaxis = dict(
        title = "Período",
        showgrid = True,
        gridcolor = "lightgrey",
        zeroline = False,
        title_font_size = 14
    ),
    yaxis = dict(
        title = "Índice de Liquidez",
        showgrid = True,
        gridcolor = "lightgrey",
        zeroline = False,
        title_font_size = 14
    ),
    width = 1000, height = 500,
    margin = dict(t = 80, b = 120)
)

formulas_text = (
    "Liquidez Corrente = Ativo Circulante ÷ Passivo Circulante<br>"
    "Liquidez Seca = (Ativo Circulante − Estoque) ÷ Passivo Circulante<br>"
    "Liquidez Imediata = Caixa e Equivalentes ÷ Passivo Circulante"
)

figure13.add_annotation(
    text = formulas_text,
    xref = "paper", yref = "paper",
    x = 0, y = -0.35,
    showarrow = False,
    font = dict(size = 12, color = "gray", family = "Arial"),
    align = "left"
)

# figure13.show()

daily_sales_quantity = database_revenue.groupby(["date", "item"]).size().reset_index(name = "quantity_sold")
daily_sales_quantity["date"] = pd.to_datetime(daily_sales_quantity["date"])

database_expense["date"] = pd.to_datetime(database_expense["date"])
daily_purchases = database_expense.groupby(["date", "material"]).agg({
    "quantity_purchased": "sum",
    "subtotal": "sum"
}).reset_index()

start_date = min(daily_sales_quantity["date"].min(), daily_purchases["date"].min())
end_date = max(daily_sales_quantity["date"].max(), daily_purchases["date"].max())
complete_date_range = pd.date_range(start = start_date, end = end_date, freq = "D")

sales_items = daily_sales_quantity["item"].unique()
purchase_materials = daily_purchases["material"].unique()
all_items = list(set(sales_items) | set(purchase_materials))

complete_panel = pd.MultiIndex.from_product(
    [complete_date_range, all_items], 
    names = ["date", "item"]
).to_frame(index = False)

complete_inventory_data = pd.merge(
    complete_panel, 
    daily_sales_quantity, 
    on = ["date", "item"], 
    how = "left"
)

daily_purchases = daily_purchases.rename(columns = {"material": "item"})
complete_inventory_data = pd.merge(
    complete_inventory_data, 
    daily_purchases[["date", "item", "quantity_purchased"]], 
    on = ["date", "item"], 
    how = "left"
)

complete_inventory_data["quantity_sold"] = complete_inventory_data["quantity_sold"].fillna(0)
complete_inventory_data["quantity_purchased"] = complete_inventory_data["quantity_purchased"].fillna(0)

complete_inventory_data["daily_net_change"] = (
    complete_inventory_data["quantity_purchased"] - complete_inventory_data["quantity_sold"]
)

complete_inventory_data = complete_inventory_data.sort_values(["item", "date"])
complete_inventory_data["cumulative_inventory_balance"] = (
    complete_inventory_data.groupby("item")["daily_net_change"].cumsum()
)

inventory_activity_summary = complete_inventory_data.groupby("item").agg({
    "quantity_sold": "sum",
    "quantity_purchased": "sum"
}).reset_index()

active_items = inventory_activity_summary[
    (inventory_activity_summary["quantity_sold"] > 0) | 
    (inventory_activity_summary["quantity_purchased"] > 0)
]["item"].unique()

filtered_inventory_data = complete_inventory_data[
    complete_inventory_data["item"].isin(active_items)
]

figure14 = px.line(
    filtered_inventory_data, 
    x = "date", 
    y = "cumulative_inventory_balance", 
    color = "item",
    labels = {
        "date": "Data", 
        "cumulative_inventory_balance": "Saldo Acumulado (Unidades)", 
        "item": "Item"
    },
    title = "Fluxo de Caixa e Estoque — Evolução Diária do Estoque por Item", 
    width = 1000, 
    height = 500
)

figure14.update_layout(
    title_font_size = 18, 
    font = dict(size = 14, family = "Arial", color = "black"),
    plot_bgcolor = "white", 
    paper_bgcolor = "white",
    legend = dict(
        title = "", 
        font_size = 12, 
        bgcolor = "rgba(0,0,0,0)"
    ),
    xaxis = dict(
        showgrid = True, 
        gridcolor = "lightgrey", 
        zeroline = False, 
        title_font_size = 14, 
        tickformat = "%d/%m/%Y"
    ),
    yaxis = dict(
        showgrid = True, 
        gridcolor = "lightgrey", 
        zeroline = False, 
        title_font_size = 14
    )
)

# figure14.show()

def obtain_npv(attractive_rate, cash_flow):
    return sum(dough / (1 + attractive_rate) ** time for time, dough in enumerate(cash_flow))

def measure_irr(cash_flow, guess = 0.1, tolerance = 1e-5):
    adjustment_rate, step = guess, 0.01

    for _ in range(1000):
        temporary_npv = obtain_npv(adjustment_rate, cash_flow)
        if abs(temporary_npv) < tolerance:
            return adjustment_rate
        adjustment_rate += step if temporary_npv > 0 else - step
        step *= 0.9

    raise ValueError("Mude o palpite!")

def discover_mirr(cash_flow, financing_rate, reinvestment_rate):
    present_value = sum(dough / (1 + financing_rate) ** time for time, dough in enumerate(cash_flow) if dough < 0)
    future_value = sum(dough * (1 + reinvestment_rate) ** (len(cash_flow) - time - 1) for time, dough in enumerate(cash_flow) if dough > 0)
    
    return (abs(future_value / present_value)) ** (1 / (len(cash_flow) - 1)) - 1

def get_selic_rate():
    url = "https://api.bcb.gov.br/dados/serie/bcdata.sgs.432/dados/ultimos/1?formato=json"
    response = requests.get(url)
    
    if response.status_code == 200:
        data = response.json()
        data = float(data[0]["valor"]) / 100
        return data
    else:
        raise Exception("Falha ao acessar a taxa Selic!")

def get_ipca_12_months():
    url = "https://api.bcb.gov.br/dados/serie/bcdata.sgs.433/dados/ultimos/12?formato=json"
    response = requests.get(url)

    if response.status_code == 200:
        data = response.json()
        monthly = [float(nibble["valor"]) / 100 for nibble in data]
        compound = 1
        for bump in monthly:
            compound *= (1 + bump)
        return compound - 1
    else:
        raise Exception("Falha ao acessar o Índice Nacional de Preços ao Consumidor Amplo (IPCA)!")

# cash_flow = [-1250, 425, 425, 425, 425, -2500, 850, 850, 850, 850]
# attractive_rate = get_selic_rate()
# financing_rate = get_selic_rate() * 1.75
# reinvestment_rate = get_ipca_12_months() + 0.05

# print(f"Valor Presente Líquido (VPL): R$ {round(obtain_npv(attractive_rate, cash_flow), 2)}")
# print(f"Taxa Interna de Retorno (TIR): {round(measure_irr(cash_flow) * 100, 2)}%")
# print(f"Taxa Interna de Retorno Modificada (TIRM): {round(discover_mirr(cash_flow, financing_rate, reinvestment_rate) * 100, 2)}%")

def buffer_csv(df):
	csv_buffer = io.StringIO()

	df.to_csv(
		csv_buffer,
		sep=";",
		decimal=",",
		index=False,
		encoding="utf-8-sig"
	)

	return csv_buffer.getvalue()


def buffer_excel_formatted(df: pd.DataFrame) -> bytes:
	excel_buffer = io.BytesIO()

	with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
		df.to_excel(writer, index=False, sheet_name="Planilha")

	excel_buffer.seek(0)

	workbook = load_workbook(excel_buffer)
	worksheet = workbook.active

	worksheet.title = "Planilha"
	worksheet.sheet_view.showGridLines = False

	header_font = Font(bold=True)

	accounting_style = NamedStyle(name="AccountingStyle")
	accounting_style.number_format = "#,##0.00"
	accounting_style.alignment = Alignment(horizontal="right", vertical="center")

	border_none = Border(
		left=Side(style=None),
		right=Side(style=None),
		top=Side(style=None),
		bottom=Side(style=None),
	)

	for cell in worksheet[1]:
		cell.font = header_font
		cell.border = border_none
		cell.alignment = Alignment(horizontal="center", vertical="center")

	for row in worksheet.iter_rows(min_row=2):
		for cell in row:
			cell.style = accounting_style
			cell.border = border_none

	for column in worksheet.columns:
		max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column)
		adjusted_width = (max_length + 2)
		worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

	excel_buffer.seek(0)
	excel_buffer.truncate()
	workbook.save(excel_buffer)

	return excel_buffer.getvalue()

with open("Script.ipynb", "r", encoding = "utf-8") as f:
    nb = nbformat.read(f, as_version = 4)

code = ""
for cell in nb.cells:
    if cell.cell_type == "code":
        code += cell.source + "\n\n"

with open("Conector.py", "w", encoding = "utf-8") as f:
    f.write(code)

